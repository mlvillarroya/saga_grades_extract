from openpyxl import Workbook 
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder 
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl import styles
from openpyxl.formula.translate import Translator

from miscellaneous.Misc import Misc

class Myexcel:
    def __init__(self,file_name,sheet_name = 'Sheet') -> None:
        self.wb = Workbook()
        self.ws = self.wb['Sheet']
        self.ws.title = sheet_name
        self.file_name = file_name

    def create_module_header_with_modules_units_hours(self,module_unit_hours_list):
        # ROW 1: MODULES, ROW 2: UNITS #
        col = 2
        for module,units_hours in module_unit_hours_list:
            start_col = col
            self.ws.cell(row=1, column=col).value = self.trim_name_until_period(module)
            self.ws.cell(row=1, column=col).comment = Comment(module,'admin')
            for unit,hours in units_hours:
                self.ws.cell(row=2, column=col).value = self.trim_name_until_period(unit)
                self.ws.cell(row=2, column=col).comment = Comment(unit,'admin')
                self.ws.cell(row=3, column=col).value = hours
                col += 1
            end_col = col - 1
            self.ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    
    def create_first_column_with_student_names(self,student_list):
        for row,student in enumerate(student_list,start=4):
            self.ws.cell(row=row, column=1).value = student
        self.adjust_col_dimensions(1)

    def adjust_col_dimensions(self,col):
        if Misc.tryparse(col): col_letter = get_column_letter(int(col)) 
        else: col_letter = col       
        dim = 0        
        for cell in self.ws[col_letter]:
            if cell.value:
                dim = max(dim, len(str(cell.value)))    
        self.ws.column_dimensions[col_letter].width = dim

    def save_file(self):
        self.wb.save(self.file_name)

    def set_passed_grades(self,row,grades):
        for col,grade in enumerate(grades,start=2):
            self.ws.cell(row=row + 4, column=col).value = self.compute_passed_grade(grade)

    def compute_passed_grade(self,grade):
        parsed_grade = Misc.tryparse(grade)
        if not parsed_grade: return 1
        elif int(grade) < 5: return 1
        else: return ''

    def adjust_columns_width(self,width,start=0,end=-1):
        if end == -1: end = self.ws.max_column
        for j in range(start,end+1):
            self.ws.column_dimensions[get_column_letter(j)].width = width

    def draw_all_lines(self):
        max_row = self.ws.max_row
        max_column = self.ws.max_column
        border_style = styles.Side(style="thin", color="000000")
        border = styles.Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                cell = self.ws.cell(row, column)
                cell.border = border

    def trim_name_until_period(self,text):
        period_position = text.find(".")
        if period_position == -1: return text
        else: return text[:period_position]

    def freeze_until(self,cell):
        self.ws.freeze_panes = cell

    def set_grades_percent(self,first_data_row,first_data_col):
        last_data_row = self.ws.max_row
        last_data_col = self.ws.max_column
        first_data_column_letter = get_column_letter(first_data_col)
        last_data_column_letter = get_column_letter(last_data_col)
        next_data_column_letter = get_column_letter(last_data_col + 1)
        formula =  "=SUMPRODUCT(" + first_data_column_letter + str(first_data_row) + ":" + last_data_column_letter + str(first_data_row) + ",$" + first_data_column_letter + "$" + str(first_data_row - 1) + ":$" + last_data_column_letter + "$" + str(first_data_row - 1) + ")/SUM($" + first_data_column_letter + "$" + str(first_data_row - 1) + ":$" + last_data_column_letter + "$" + str(first_data_row - 1) + ")"
        source_cell = self.ws[next_data_column_letter + str(first_data_row)]
        source_cell.value = formula
        self.change_to_percentage_column(next_data_column_letter,first_data_row,last_data_row)
        self.extend_formula_in_column(formula,next_data_column_letter,first_data_row,last_data_row)
        self.adjust_columns_width(8,last_data_col+1)

    def extend_formula_in_column(self,formula,col_letter : str,first_data_row,last_data_row):
        for i in range(first_data_row + 1, last_data_row + 1): # Rango de columnas de K a Y
            source_cell = self.ws[col_letter + str(first_data_row)]
            end_cell = self.ws[col_letter + str(i)]
            translator = Translator(formula, source_cell.coordinate) 
            new_formula = translator.translate_formula(end_cell.coordinate) 
            end_cell.value = new_formula 

    def change_to_percentage_column(self,column_letter : str,first_data_row,last_data_row):
        for i in range(first_data_row,last_data_row + 1):
            cell = self.ws[column_letter + str(i)]
            cell.number_format = '0.00%' 

    def create_passed_units_statistics(self,students_list,module_unit_names_hours,unit_grades_list):
        self.create_module_header_with_modules_units_hours(module_unit_names_hours)
        self.create_first_column_with_student_names(students_list)
        for i,student in enumerate(students_list):
            self.set_passed_grades(i,unit_grades_list[i])
        self.adjust_columns_width(5,2)
        self.set_grades_percent(4,2)
        self.draw_all_lines()
        self.freeze_until('B4')
        self.save_file()